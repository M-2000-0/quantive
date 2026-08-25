"""Excel export utilities using openpyxl for portfolios, optimizations, and reports."""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
MONEY_FORMAT = '#,##0'
PCT_FORMAT = '0.00%'
DATE_FORMAT = 'YYYY-MM-DD'


def _style_header(ws, headers: list[str]):
    """Apply header styling to the first row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _auto_width(ws):
    """Auto-fit column widths."""
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, min(ws.max_row + 1, 50)):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 40)


def _apply_borders(ws):
    """Apply borders to all data cells."""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER


def export_portfolio_excel(portfolio: dict) -> BytesIO:
    """Export portfolio with instruments to Excel.

    Args:
        portfolio: dict with keys: name, description, instruments (list of dicts)
    """
    wb = Workbook()

    # ── Summary Sheet ─────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.sheet_properties.tabColor = "1E40AF"

    summary_data = [
        ["Portfolio Name", portfolio.get("name", "")],
        ["Description", portfolio.get("description", "")],
        ["Total Instruments", len(portfolio.get("instruments", []))],
        ["Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")],
    ]

    for row_idx, (label, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True, size=11)
        ws_summary.cell(row=row_idx, column=2, value=value)

    total_principal = sum(
        inst.get("principal_outstanding", 0) for inst in portfolio.get("instruments", [])
    )
    ws_summary.cell(row=6, column=1, value="Total Principal").font = Font(bold=True, size=11)
    ws_summary.cell(row=6, column=2, value=total_principal).number_format = MONEY_FORMAT

    currencies = set(inst.get("currency", "") for inst in portfolio.get("instruments", []))
    ws_summary.cell(row=7, column=1, value="Currencies").font = Font(bold=True, size=11)
    ws_summary.cell(row=7, column=2, value=", ".join(sorted(currencies)))

    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 50

    # ── Instruments Sheet ─────────────────────────────────────────────
    ws_inst = wb.create_sheet("Instruments")
    ws_inst.sheet_properties.tabColor = "10B981"

    headers = [
        "Name", "Type", "Currency", "Principal Outstanding",
        "Coupon Rate", "Maturity Date", "Issue Date",
        "Spread (bps)", "Callable", "Call Date", "Call Price",
    ]
    _style_header(ws_inst, headers)

    for row_idx, inst in enumerate(portfolio.get("instruments", []), 2):
        ws_inst.cell(row=row_idx, column=1, value=inst.get("name", ""))
        ws_inst.cell(row=row_idx, column=2, value=inst.get("instrument_type", ""))
        ws_inst.cell(row=row_idx, column=3, value=inst.get("currency", ""))

        principal_cell = ws_inst.cell(row=row_idx, column=4, value=inst.get("principal_outstanding", 0))
        principal_cell.number_format = MONEY_FORMAT

        coupon_cell = ws_inst.cell(row=row_idx, column=5, value=inst.get("coupon_rate", 0))
        coupon_cell.number_format = PCT_FORMAT

        ws_inst.cell(row=row_idx, column=6, value=inst.get("maturity_date", ""))
        ws_inst.cell(row=row_idx, column=7, value=inst.get("issue_date", ""))
        ws_inst.cell(row=row_idx, column=8, value=inst.get("spread_bps", 0))
        ws_inst.cell(row=row_idx, column=9, value="Yes" if inst.get("is_callable") else "No")
        ws_inst.cell(row=row_idx, column=10, value=inst.get("call_date") or "")
        call_price = inst.get("call_price")
        ws_inst.cell(row=row_idx, column=11, value=call_price if call_price else "")

    _apply_borders(ws_inst)
    _auto_width(ws_inst)

    # ── Currency Breakdown Sheet ──────────────────────────────────────
    ws_ccy = wb.create_sheet("By Currency")
    ws_ccy.sheet_properties.tabColor = "F59E0B"

    ccy_headers = ["Currency", "Count", "Total Principal", "% of Portfolio"]
    _style_header(ws_ccy, ccy_headers)

    ccy_data: dict[str, dict] = {}
    for inst in portfolio.get("instruments", []):
        ccy = inst.get("currency", "UNKNOWN")
        if ccy not in ccy_data:
            ccy_data[ccy] = {"count": 0, "total": 0}
        ccy_data[ccy]["count"] += 1
        ccy_data[ccy]["total"] += inst.get("principal_outstanding", 0)

    for row_idx, (ccy, data) in enumerate(sorted(ccy_data.items()), 2):
        ws_ccy.cell(row=row_idx, column=1, value=ccy)
        ws_ccy.cell(row=row_idx, column=2, value=data["count"])
        ws_ccy.cell(row=row_idx, column=3, value=data["total"]).number_format = MONEY_FORMAT
        pct = data["total"] / total_principal if total_principal > 0 else 0
        ws_ccy.cell(row=row_idx, column=4, value=pct).number_format = PCT_FORMAT

    _apply_borders(ws_ccy)
    _auto_width(ws_ccy)

    # ── Save ──────────────────────────────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_optimization_excel(job: dict, strategies: list[dict], benchmarks: list[dict]) -> BytesIO:
    """Export optimization results to Excel."""
    wb = Workbook()

    # ── Summary Sheet ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "1E40AF"

    summary = [
        ["Optimization Name", job.get("name", "")],
        ["Status", job.get("status", "")],
        ["Type", job.get("optimization_type", "")],
        ["Random Seed", job.get("random_seed", "")],
        ["Created", job.get("created_at", "")],
        ["Completed", job.get("completed_at", "N/A")],
    ]
    for row_idx, (label, value) in enumerate(summary, 1):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True, size=11)
        ws.cell(row=row_idx, column=2, value=str(value))
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 50

    # ── Strategies Sheet ──────────────────────────────────────────────
    if strategies:
        ws_strat = wb.create_sheet("Strategies")
        ws_strat.sheet_properties.tabColor = "10B981"

        strat_headers = ["Rank", "Name", "Description", "Expected Cost", "Refinancing Risk", "Interest Rate Risk", "Currency Risk"]
        _style_header(ws_strat, strat_headers)

        for row_idx, s in enumerate(strategies, 2):
            ws_strat.cell(row=row_idx, column=1, value=s.get("rank", row_idx - 1))
            ws_strat.cell(row=row_idx, column=2, value=s.get("name", ""))
            ws_strat.cell(row=row_idx, column=3, value=s.get("description", ""))
            metrics = s.get("metrics", {})
            ws_strat.cell(row=row_idx, column=4, value=metrics.get("expected_cost", 0)).number_format = MONEY_FORMAT
            ws_strat.cell(row=row_idx, column=5, value=metrics.get("refinancing_risk", 0)).number_format = PCT_FORMAT
            ws_strat.cell(row=row_idx, column=6, value=metrics.get("interest_rate_risk", 0)).number_format = PCT_FORMAT
            ws_strat.cell(row=row_idx, column=7, value=metrics.get("currency_risk", 0)).number_format = PCT_FORMAT

        _apply_borders(ws_strat)
        _auto_width(ws_strat)

    # ── Benchmarks Sheet ──────────────────────────────────────────────
    if benchmarks:
        ws_bench = wb.create_sheet("Benchmarks")
        ws_bench.sheet_properties.tabColor = "8B5CF6"

        bench_headers = ["Solver", "Feasible", "Objective Value", "Runtime (s)", "Iterations", "Compute Cost"]
        _style_header(ws_bench, bench_headers)

        for row_idx, b in enumerate(benchmarks, 2):
            ws_bench.cell(row=row_idx, column=1, value=b.get("solver_name", ""))
            ws_bench.cell(row=row_idx, column=2, value="Yes" if b.get("feasible") else "No")
            ws_bench.cell(row=row_idx, column=3, value=b.get("objective_value", 0)).number_format = MONEY_FORMAT
            ws_bench.cell(row=row_idx, column=4, value=b.get("execution_time_seconds", 0)).number_format = '0.0'
            ws_bench.cell(row=row_idx, column=5, value=b.get("iterations", 0))
            metrics = b.get("metrics", {})
            ws_bench.cell(row=row_idx, column=6, value=metrics.get("compute_cost", 0)).number_format = '0.00'

        _apply_borders(ws_bench)
        _auto_width(ws_bench)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_risk_excel(risk_summary: dict) -> BytesIO:
    """Export risk analysis results to Excel."""
    wb = Workbook()

    # ── Investment Scenarios ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Investment Scenarios"
    ws.sheet_properties.tabColor = "EF4444"

    headers = ["Scenario", "Investment", "Return Amount", "Return %", "Probability", "Description"]
    _style_header(ws, headers)

    for row_idx, sc in enumerate(risk_summary.get("investment_scenarios", []), 2):
        ws.cell(row=row_idx, column=1, value=sc.get("scenario_name", ""))
        ws.cell(row=row_idx, column=2, value=sc.get("investment", 0)).number_format = MONEY_FORMAT
        ws.cell(row=row_idx, column=3, value=sc.get("return_amount", 0)).number_format = MONEY_FORMAT
        ws.cell(row=row_idx, column=4, value=sc.get("return_pct", 0) / 100).number_format = PCT_FORMAT
        ws.cell(row=row_idx, column=5, value=sc.get("probability", 0)).number_format = PCT_FORMAT
        ws.cell(row=row_idx, column=6, value=sc.get("description", ""))

    _apply_borders(ws)
    _auto_width(ws)

    # ── Risk Score ────────────────────────────────────────────────────
    risk_score = risk_summary.get("risk_score", {})
    if risk_score:
        ws_score = wb.create_sheet("Risk Score")
        ws_score.sheet_properties.tabColor = "F59E0B"

        ws_score.cell(row=1, column=1, value="Overall Risk Score").font = Font(bold=True, size=12)
        ws_score.cell(row=1, column=2, value=risk_score.get("score", 0))
        ws_score.cell(row=2, column=1, value="Label").font = Font(bold=True)
        ws_score.cell(row=2, column=2, value=risk_score.get("label", ""))

        comp_headers = ["Component", "Score", "Weight", "Description"]
        _style_header(ws_score, comp_headers)

        for row_idx, (key, val) in enumerate(risk_score.get("components", {}).items(), 4):
            ws_score.cell(row=row_idx, column=1, value=key.replace("_", " ").title())
            ws_score.cell(row=row_idx, column=2, value=val.get("score", 0))
            ws_score.cell(row=row_idx, column=3, value=val.get("weight", 0)).number_format = PCT_FORMAT
            ws_score.cell(row=row_idx, column=4, value=val.get("description", ""))

        _apply_borders(ws_score)
        _auto_width(ws_score)

    # ── VaR Analysis ──────────────────────────────────────────────────
    var_results = risk_summary.get("var_analysis", [])
    if var_results:
        ws_var = wb.create_sheet("VaR Analysis")
        ws_var.sheet_properties.tabColor = "DC2626"

        var_headers = ["Confidence", "Horizon (days)", "VaR Amount", "VaR %", "CVaR Amount", "CVaR %"]
        _style_header(ws_var, var_headers)

        for row_idx, v in enumerate(var_results, 2):
            ws_var.cell(row=row_idx, column=1, value=f"{v.get('confidence', 0) * 100:.0f}%")
            ws_var.cell(row=row_idx, column=2, value=v.get("horizon_days", 0))
            ws_var.cell(row=row_idx, column=3, value=v.get("var_amount", 0)).number_format = MONEY_FORMAT
            ws_var.cell(row=row_idx, column=4, value=v.get("var_pct", 0)).number_format = PCT_FORMAT
            ws_var.cell(row=row_idx, column=5, value=v.get("cvar_amount", 0)).number_format = MONEY_FORMAT
            ws_var.cell(row=row_idx, column=6, value=v.get("cvar_pct", 0)).number_format = PCT_FORMAT

        _apply_borders(ws_var)
        _auto_width(ws_var)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
