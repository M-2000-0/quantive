"""Excel Import for Portfolios.

Parses uploaded Excel files into instrument data for portfolio creation.
Supports multiple formats and auto-detects column mappings.
"""

from __future__ import annotations

from datetime import datetime
from typing import Optional

from openpyxl import load_workbook

# Column name mappings (flexible matching)
COLUMN_ALIASES = {
    "name": ["name", "instrument", "instrument name", "bond name", "security", "description"],
    "instrument_type": ["type", "instrument_type", "instrument type", "bond type", "category"],
    "currency": ["currency", "ccy", "denomination"],
    "principal_outstanding": ["principal", "amount", "outstanding", "notional", "face value", "par", "principal_outstanding"],
    "coupon_rate": ["coupon", "rate", "coupon_rate", "coupon rate", "interest rate"],
    "maturity_date": ["maturity", "maturity_date", "maturity date", "maturity_date", "expiry", "due date"],
    "issue_date": ["issue", "issue_date", "issue date", "settlement", "settlement date"],
    "spread_bps": ["spread", "spread_bps", "spread (bps)", "credit spread"],
    "is_callable": ["callable", "is_callable", "call option"],
    "call_date": ["call_date", "call date", "first call"],
    "call_price": ["call_price", "call price"],
}

TYPE_ALIASES = {
    "treasury_bond": ["treasury", "treasury bond", "govt bond", "government bond", "t-bond"],
    "treasury_bill": ["t-bill", "treasury bill", "tbill"],
    "t_bill": ["t-bill", "treasury bill", "tbill"],
    "sovereign_bond": ["sovereign", "sovereign bond", "sovereign debt"],
    "domestic_bond": ["domestic", "domestic bond", "local bond"],
    "eurobond": ["eurobond", "euro bond", "external bond"],
    "floating_rate_note": ["frn", "floating", "floating rate", "floating rate note", "floater"],
    "inflation_linked": ["inflation", "inflation-linked", "ilb", "tips"],
    "concessional_loan": ["concessional", "concessional loan", "multilateral loan"],
    "commercial_loan": ["commercial", "commercial loan", "syndicated loan"],
}

CURRENCY_ALIASES = {
    "USD": ["usd", "$", "us dollar", "dollar"],
    "EUR": ["eur", "€", "euro"],
    "GBP": ["gbp", "£", "pound", "sterling"],
    "JPY": ["jpy", "¥", "yen"],
    "CHF": ["chf", "franc"],
    "CAD": ["cad", "canadian dollar"],
    "AUD": ["aud", "australian dollar"],
    "CNY": ["cny", "rmb", "yuan", "chinese yuan"],
    "INR": ["inr", "rupee", "indian rupee"],
    "BRL": ["brl", "real", "brazilian real"],
}


def parse_excel_portfolio(file_content: bytes, filename: str = "") -> dict:
    """Parse an Excel file into portfolio data.

    Args:
        file_content: Raw bytes of the Excel file.
        filename: Original filename for context.

    Returns:
        {
            "name": "Portfolio from {filename}",
            "instruments": [...],
            "warnings": [...],
            "stats": {...}
        }
    """
    wb = load_workbook(file_content, read_only=True, data_only=True)

    # Find the best sheet (first sheet with data)
    ws = None
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if sheet.max_row and sheet.max_row > 2:
            ws = sheet
            break

    if not ws:
        return {"name": f"Portfolio from {filename}", "instruments": [], "warnings": ["No data found in file"], "stats": {}}

    # Read headers
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
        headers.append(str(cell).strip().lower() if cell else "")

    # Map columns
    col_map = _map_columns(headers)

    # Parse instruments
    instruments = []
    warnings = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        try:
            instrument = _parse_row(row, col_map, headers)
            if instrument and instrument.get("name"):
                instruments.append(instrument)
        except Exception as e:
            warnings.append(f"Row {row_idx}: {str(e)}")

    # Derive portfolio name from filename
    portfolio_name = filename.rsplit(".", 1)[0].replace("_", " ").replace("-", " ").title() if filename else "Imported Portfolio"

    wb.close()

    return {
        "name": portfolio_name,
        "instruments": instruments,
        "warnings": warnings,
        "stats": {
            "total_instruments": len(instruments),
            "currencies": list(set(i.get("currency", "USD") for i in instruments)),
            "total_principal": sum(i.get("principal_outstanding", 0) for i in instruments),
            "columns_mapped": len([v for v in col_map.values() if v is not None]),
            "total_columns": len(headers),
        },
    }


def _map_columns(headers: list[str]) -> dict:
    """Map Excel columns to instrument fields."""
    col_map = {}

    for field, aliases in COLUMN_ALIASES.items():
        for idx, header in enumerate(headers):
            header_lower = header.lower().strip()
            if header_lower in aliases or any(alias in header_lower for alias in aliases):
                col_map[field] = idx
                break

    return col_map


def _parse_row(row: tuple, col_map: dict, headers: list[str]) -> Optional[dict]:
    """Parse a single Excel row into an instrument dict."""
    instrument = {}

    # Name (required)
    name_idx = col_map.get("name")
    if name_idx is not None and name_idx < len(row):
        name = str(row[name_idx] or "").strip()
        if not name:
            return None
        instrument["name"] = name
    else:
        return None

    # Instrument type
    type_idx = col_map.get("instrument_type")
    if type_idx is not None and type_idx < len(row):
        raw_type = str(row[type_idx] or "").strip().lower()
        instrument["instrument_type"] = _resolve_type(raw_type)
    else:
        instrument["instrument_type"] = "treasury_bond"

    # Currency
    ccy_idx = col_map.get("currency")
    if ccy_idx is not None and ccy_idx < len(row):
        raw_ccy = str(row[ccy_idx] or "").strip().upper()
        instrument["currency"] = _resolve_currency(raw_ccy)
    else:
        instrument["currency"] = "USD"

    # Principal
    principal_idx = col_map.get("principal_outstanding")
    if principal_idx is not None and principal_idx < len(row):
        val = row[principal_idx]
        if val is not None:
            instrument["principal_outstanding"] = _parse_number(val)
        else:
            instrument["principal_outstanding"] = 0
    else:
        instrument["principal_outstanding"] = 0

    # Coupon rate
    coupon_idx = col_map.get("coupon_rate")
    if coupon_idx is not None and coupon_idx < len(row):
        val = row[coupon_idx]
        if val is not None:
            rate = _parse_number(val)
            # Auto-detect if percentage or decimal
            if rate > 1:
                rate = rate / 100
            instrument["coupon_rate"] = rate
        else:
            instrument["coupon_rate"] = 0
    else:
        instrument["coupon_rate"] = 0

    # Maturity date
    mat_idx = col_map.get("maturity_date")
    if mat_idx is not None and mat_idx < len(row):
        instrument["maturity_date"] = _parse_date(row[mat_idx])
    else:
        instrument["maturity_date"] = "2030-01-01"

    # Issue date
    issue_idx = col_map.get("issue_date")
    if issue_idx is not None and issue_idx < len(row):
        instrument["issue_date"] = _parse_date(row[issue_idx])
    else:
        instrument["issue_date"] = "2025-01-01"

    # Spread
    spread_idx = col_map.get("spread_bps")
    if spread_idx is not None and spread_idx < len(row):
        val = row[spread_idx]
        instrument["spread_bps"] = _parse_number(val) if val else 0
    else:
        instrument["spread_bps"] = 0

    # Callable
    callable_idx = col_map.get("is_callable")
    if callable_idx is not None and callable_idx < len(row):
        val = str(row[callable_idx] or "").strip().lower()
        instrument["is_callable"] = val in ("yes", "true", "1", "callable")
    else:
        instrument["is_callable"] = False

    # Call date
    call_date_idx = col_map.get("call_date")
    if call_date_idx is not None and call_date_idx < len(row):
        instrument["call_date"] = _parse_date(row[call_date_idx])
    else:
        instrument["call_date"] = None

    # Call price
    call_price_idx = col_map.get("call_price")
    if call_price_idx is not None and call_price_idx < len(row):
        val = row[call_price_idx]
        instrument["call_price"] = _parse_number(val) if val else None
    else:
        instrument["call_price"] = None

    return instrument


def _resolve_type(raw: str) -> str:
    """Resolve instrument type from raw text."""
    for canonical, aliases in TYPE_ALIASES.items():
        if raw in aliases or any(alias in raw for alias in aliases):
            return canonical
    return "treasury_bond"  # default


def _resolve_currency(raw: str) -> str:
    """Resolve currency code from raw text."""
    for code, aliases in CURRENCY_ALIASES.items():
        if raw == code or raw in aliases:
            return code
    if len(raw) == 3 and raw.isalpha():
        return raw.upper()
    return "USD"


def _parse_number(val) -> float:
    """Parse a numeric value from various formats."""
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        cleaned = val.replace(",", "").replace("$", "").replace("€", "").replace("£", "").strip()
        try:
            return float(cleaned)
        except ValueError:
            return 0
    return 0


def _parse_date(val) -> str:
    """Parse a date value into YYYY-MM-DD format."""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        # Try common formats
        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d %b %Y", "%d %B %Y"]:
            try:
                return datetime.strptime(val.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        # Try Excel serial date
        try:
            num = float(val)
            if num > 40000:  # Excel serial date range
                from datetime import timedelta
                base = datetime(1899, 12, 30)
                return (base + timedelta(days=int(num))).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return "2030-01-01"


def generate_import_template() -> bytes:
    """Generate an empty Excel template for portfolio import.

    Returns:
        Bytes of an .xlsx file with headers and examples.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio"

    # Headers
    headers = [
        "Name", "Type", "Currency", "Principal Outstanding",
        "Coupon Rate", "Maturity Date", "Issue Date",
        "Spread (bps)", "Callable", "Call Date", "Call Price",
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Example rows
    examples = [
        ["US Treasury 10Y Bond", "treasury_bond", "USD", 12500000000, 0.0425, "2035-06-15", "2025-06-15", 0, "No", "", ""],
        ["US Treasury 5Y Bond", "treasury_bond", "USD", 8000000000, 0.0375, "2031-03-15", "2026-03-15", 0, "No", "", ""],
        ["EUR Sovereign Bond", "sovereign_bond", "EUR", 5000000000, 0.028, "2033-04-15", "2023-04-15", 12, "No", "", ""],
        ["Floating Rate Note", "floating_rate_note", "USD", 3500000000, 0.0045, "2028-12-01", "2025-12-01", 15, "No", "", ""],
        ["T-Bill 3-Month", "t_bill", "USD", 4500000000, 0.0, "2026-11-15", "2026-08-15", 0, "No", "", ""],
    ]

    for row_idx, example in enumerate(examples, 2):
        for col, val in enumerate(example, 1):
            ws.cell(row=row_idx, column=col, value=val)

    # Auto-width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = max(len(headers[col-1]) + 4, 15)

    output = __import__("io").BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
